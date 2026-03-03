from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Count, Sum, Avg, Q, Case, When, FloatField
from django.db.models.functions import TruncMonth, TruncDay
from django.utils import timezone
from datetime import timedelta, datetime
import csv
import io
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from students.models import Student
from teachers.models import Teacher
from academics.models import Class, Subject, Section
from attendance.models import StudentAttendance
from exams.models import Exam, ExamResult
from finance.models import Payment, Expense, FeeAssignment
from library.models import Book, BookIssue
from communication.models import Announcement
from accounts.models import AuditLog
from core.reports.serializers import ExportFormatSerializer

class DashboardReportView(APIView):
    """
    Get comprehensive dashboard statistics
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        today = timezone.now().date()
        first_day_month = today.replace(day=1)
        
        # Basic counts
        total_students = Student.objects.filter(is_active=True).count()
        total_teachers = Teacher.objects.filter(is_active=True).count()
        total_classes = Class.objects.filter(is_active=True).count()
        
        # Today's attendance
        today_attendance = StudentAttendance.objects.filter(date=today)
        today_present = today_attendance.filter(status='present').count()
        today_absent = today_attendance.filter(status='absent').count()
        today_late = today_attendance.filter(status='late').count()
        
        # Monthly finance
        monthly_payments = Payment.objects.filter(
            payment_date__gte=first_day_month,
            status='completed'
        )
        monthly_revenue = monthly_payments.aggregate(total=Sum('amount'))['total'] or 0
        
        monthly_expenses = Expense.objects.filter(
            expense_date__gte=first_day_month
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Pending fees
        pending_fees = FeeAssignment.objects.filter(
            status__in=['pending', 'partial', 'overdue']
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Upcoming events/announcements
        upcoming_announcements = Announcement.objects.filter(
            publish_from__gte=today,
            is_active=True
        ).count()
        
        # Recent activities
        recent_activities = AuditLog.objects.select_related('user').order_by('-timestamp')[:10]
        activities_data = []
        for log in recent_activities:
            activities_data.append({
                'user': log.user.full_name if log.user else 'System',
                'action': log.action,
                'model': log.model_name,
                'timestamp': log.timestamp
            })
        
        # Attendance chart data (last 7 days)
        last_week = today - timedelta(days=7)
        attendance_chart = StudentAttendance.objects.filter(
            date__gte=last_week
        ).values('date').annotate(
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late'))
        ).order_by('date')
        
        # Revenue chart data (last 6 months)
        revenue_chart = []
        for i in range(5, -1, -1):
            month = today - timedelta(days=30*i)
            month_start = month.replace(day=1)
            if month.month == 12:
                month_end = month.replace(day=31)
            else:
                month_end = month.replace(month=month.month+1, day=1) - timedelta(days=1)
            
            revenue = Payment.objects.filter(
                payment_date__range=[month_start, month_end],
                status='completed'
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            revenue_chart.append({
                'month': month.strftime('%B %Y'),
                'revenue': float(revenue)
            })
        
        return Response({
            'stats': {
                'total_students': total_students,
                'total_teachers': total_teachers,
                'total_classes': total_classes,
                'today_attendance': {
                    'present': today_present,
                    'absent': today_absent,
                    'late': today_late,
                    'total': today_present + today_absent + today_late
                },
                'monthly_finance': {
                    'revenue': monthly_revenue,
                    'expenses': monthly_expenses,
                    'net': monthly_revenue - monthly_expenses
                },
                'pending_fees': pending_fees,
                'upcoming_events': upcoming_announcements
            },
            'charts': {
                'attendance': list(attendance_chart),
                'revenue': revenue_chart
            },
            'recent_activities': activities_data
        })

class StudentReportView(APIView):
    """
    Generate student reports
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        report_type = request.query_params.get('type', 'summary')
        
        if report_type == 'summary':
            return self.get_summary_report(request)
        elif report_type == 'class_wise':
            return self.get_class_wise_report(request)
        elif report_type == 'gender':
            return self.get_gender_report(request)
        elif report_type == 'detailed':
            return self.get_detailed_report(request)
        else:
            return Response({'error': 'Invalid report type'}, status=400)
    
    def get_summary_report(self, request):
        total = Student.objects.count()
        active = Student.objects.filter(is_active=True).count()
        alumni = Student.objects.filter(is_alumni=True).count()
        
        # New admissions this month
        month_start = timezone.now().replace(day=1)
        new_admissions = Student.objects.filter(
            admission_date__gte=month_start
        ).count()
        
        return Response({
            'total_students': total,
            'active_students': active,
            'alumni': alumni,
            'new_admissions_this_month': new_admissions
        })
    
    def get_class_wise_report(self, request):
        report = []
        classes = Class.objects.filter(is_active=True)
        
        for cls in classes:
            total = Student.objects.filter(current_class=cls, is_active=True).count()
            boys = Student.objects.filter(
                current_class=cls, is_active=True, gender='M'
            ).count()
            girls = Student.objects.filter(
                current_class=cls, is_active=True, gender='F'
            ).count()
            
            sections = []
            for section in cls.sections.filter(is_active=True):
                section_count = Student.objects.filter(
                    current_class=cls,
                    current_section=section,
                    is_active=True
                ).count()
                sections.append({
                    'name': section.name,
                    'count': section_count,
                    'capacity': section.capacity,
                    'available': section.capacity - section_count
                })
            
            report.append({
                'class': cls.name,
                'total': total,
                'boys': boys,
                'girls': girls,
                'sections': sections
            })
        
        return Response(report)
    
    def get_gender_report(self, request):
        boys = Student.objects.filter(gender='M').count()
        girls = Student.objects.filter(gender='F').count()
        other = Student.objects.filter(gender='O').count()
        
        return Response({
            'boys': boys,
            'girls': girls,
            'other': other,
            'total': boys + girls + other
        })
    
    def get_detailed_report(self, request):
        class_id = request.query_params.get('class_id')
        section_id = request.query_params.get('section_id')
        
        students = Student.objects.filter(is_active=True).select_related(
            'user', 'current_class', 'current_section'
        )
        
        if class_id:
            students = students.filter(current_class_id=class_id)
        if section_id:
            students = students.filter(current_section_id=section_id)
        
        data = []
        for student in students.order_by('current_class__name', 'roll_number'):
            data.append({
                'admission_number': student.admission_number,
                'roll_number': student.roll_number,
                'name': student.user.full_name,
                'class': student.current_class.name if student.current_class else None,
                'section': student.current_section.name if student.current_section else None,
                'gender': student.get_gender_display(),
                'date_of_birth': student.date_of_birth,
                'blood_group': student.blood_group,
                'address': f"{student.address_line1}, {student.city}, {student.state}",
                'parents': [
                    {'name': p.user.full_name, 'relation': p.relationship}
                    for p in student.parents.all()
                ]
            })
        
        return Response(data)

class FinanceReportView(APIView):
    """
    Generate financial reports
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        report_type = request.query_params.get('type', 'summary')
        
        if report_type == 'summary':
            return self.get_summary_report(request)
        elif report_type == 'monthly':
            return self.get_monthly_report(request)
        elif report_type == 'fee_collection':
            return self.get_fee_collection_report(request)
        elif report_type == 'expense':
            return self.get_expense_report(request)
        else:
            return Response({'error': 'Invalid report type'}, status=400)
    
    def get_summary_report(self, request):
        year = request.query_params.get('year', timezone.now().year)
        
        # Total revenue
        total_revenue = Payment.objects.filter(
            payment_date__year=year,
            status='completed'
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Total expenses
        total_expenses = Expense.objects.filter(
            expense_date__year=year
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Pending fees
        pending_fees = FeeAssignment.objects.filter(
            status__in=['pending', 'partial', 'overdue']
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Monthly breakdown
        monthly_data = []
        for month in range(1, 13):
            revenue = Payment.objects.filter(
                payment_date__year=year,
                payment_date__month=month,
                status='completed'
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            expenses = Expense.objects.filter(
                expense_date__year=year,
                expense_date__month=month
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            monthly_data.append({
                'month': datetime(year, month, 1).strftime('%B'),
                'revenue': revenue,
                'expenses': expenses,
                'net': revenue - expenses
            })
        
        return Response({
            'year': year,
            'total_revenue': total_revenue,
            'total_expenses': total_expenses,
            'net_income': total_revenue - total_expenses,
            'pending_fees': pending_fees,
            'monthly_data': monthly_data
        })
    
    def get_monthly_report(self, request):
        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))
        
        # Daily revenue for the month
        daily_revenue = Payment.objects.filter(
            payment_date__year=year,
            payment_date__month=month,
            status='completed'
        ).annotate(
            day=TruncDay('payment_date')
        ).values('day').annotate(
            total=Sum('amount')
        ).order_by('day')
        
        # Daily expenses
        daily_expenses = Expense.objects.filter(
            expense_date__year=year,
            expense_date__month=month
        ).annotate(
            day=TruncDay('expense_date')
        ).values('day').annotate(
            total=Sum('amount')
        ).order_by('day')
        
        # Payment methods breakdown
        payment_methods = Payment.objects.filter(
            payment_date__year=year,
            payment_date__month=month,
            status='completed'
        ).values('payment_method').annotate(
            total=Sum('amount'),
            count=Count('id')
        )
        
        # Expense categories
        expense_categories = Expense.objects.filter(
            expense_date__year=year,
            expense_date__month=month
        ).values('category').annotate(
            total=Sum('amount'),
            count=Count('id')
        )
        
        return Response({
            'month': datetime(year, month, 1).strftime('%B %Y'),
            'daily_revenue': daily_revenue,
            'daily_expenses': daily_expenses,
            'payment_methods': payment_methods,
            'expense_categories': expense_categories
        })
    
    def get_fee_collection_report(self, request):
        class_id = request.query_params.get('class_id')
        
        assignments = FeeAssignment.objects.all()
        if class_id:
            assignments = assignments.filter(student__current_class_id=class_id)
        
        # Collection by fee type
        by_fee_type = assignments.values(
            'fee_structure__fee_type',
            'fee_structure__name'
        ).annotate(
            total_assigned=Sum('amount'),
            collected=Sum('payments__amount', filter=Q(payments__status='completed')),
            pending=Sum('amount') - Sum('payments__amount', filter=Q(payments__status='completed')),
            student_count=Count('student', distinct=True)
        )
        
        # Collection by class
        by_class = assignments.values(
            'student__current_class__name'
        ).annotate(
            total_assigned=Sum('amount'),
            collected=Sum('payments__amount', filter=Q(payments__status='completed')),
            pending=Sum('amount') - Sum('payments__amount', filter=Q(payments__status='completed')),
            student_count=Count('student', distinct=True)
        )
        
        # Top defaulters
        defaulters = assignments.filter(
            status='overdue'
        ).values(
            'student__user__first_name',
            'student__user__last_name',
            'student__admission_number',
            'student__current_class__name'
        ).annotate(
            total_due=Sum('amount'),
            paid=Sum('payments__amount', filter=Q(payments__status='completed')),
            balance=Sum('amount') - Sum('payments__amount', filter=Q(payments__status='completed'))
        ).order_by('-balance')[:20]
        
        return Response({
            'by_fee_type': by_fee_type,
            'by_class': by_class,
            'top_defaulters': defaulters
        })
    
    def get_expense_report(self, request):
        year = request.query_params.get('year', timezone.now().year)
        
        # Expenses by category
        by_category = Expense.objects.filter(
            expense_date__year=year
        ).values('category').annotate(
            total=Sum('amount'),
            count=Count('id'),
            average=Avg('amount')
        ).order_by('-total')
        
        # Monthly expense trend
        monthly = Expense.objects.filter(
            expense_date__year=year
        ).annotate(
            month=TruncMonth('expense_date')
        ).values('month').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('month')
        
        # Largest expenses
        largest = Expense.objects.filter(
            expense_date__year=year
        ).order_by('-amount')[:10]
        
        largest_data = []
        for expense in largest:
            largest_data.append({
                'date': expense.expense_date,
                'category': expense.get_category_display(),
                'description': expense.description,
                'amount': expense.amount,
                'vendor': expense.vendor
            })
        
        return Response({
            'year': year,
            'by_category': by_category,
            'monthly_trend': monthly,
            'largest_expenses': largest_data
        })

class AttendanceReportView(APIView):
    """
    Generate attendance reports
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        report_type = request.query_params.get('type', 'summary')
        
        if report_type == 'summary':
            return self.get_summary_report(request)
        elif report_type == 'student':
            return self.get_student_report(request)
        elif report_type == 'class':
            return self.get_class_report(request)
        elif report_type == 'monthly':
            return self.get_monthly_report(request)
        else:
            return Response({'error': 'Invalid report type'}, status=400)
    
    def get_summary_report(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not start_date:
            start_date = timezone.now().date().replace(day=1)
        if not end_date:
            end_date = timezone.now().date()
        
        attendances = StudentAttendance.objects.filter(
            date__range=[start_date, end_date]
        )
        
        total_days = attendances.values('date').distinct().count()
        total_records = attendances.count()
        
        summary = attendances.values('status').annotate(
            count=Count('id')
        )
        
        # Daily breakdown
        daily = attendances.values('date').annotate(
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
            half_day=Count('id', filter=Q(status='half_day'))
        ).order_by('date')
        
        # Class-wise breakdown
        class_wise = attendances.values(
            'class_group__name',
            'section__name'
        ).annotate(
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            total=Count('id')
        )
        
        return Response({
            'period': {'start': start_date, 'end': end_date},
            'total_days': total_days,
            'total_records': total_records,
            'summary': summary,
            'daily_breakdown': daily,
            'class_wise': class_wise
        })
    
    def get_student_report(self, request):
        student_id = request.query_params.get('student_id')
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        
        if not student_id:
            return Response({'error': 'student_id required'}, status=400)
        
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=404)
        
        attendances = StudentAttendance.objects.filter(student=student)
        
        if month and year:
            attendances = attendances.filter(
                date__year=year,
                date__month=month
            )
        
        total_days = attendances.count()
        present = attendances.filter(status='present').count()
        absent = attendances.filter(status='absent').count()
        late = attendances.filter(status='late').count()
        half_day = attendances.filter(status='half_day').count()
        
        attendance_percentage = ((present + late + (half_day * 0.5)) / total_days * 100) if total_days > 0 else 0
        
        daily_records = attendances.order_by('date').values('date', 'status')
        
        return Response({
            'student': {
                'id': student.id,
                'name': student.user.full_name,
                'class': student.current_class.name if student.current_class else None,
                'section': student.current_section.name if student.current_section else None,
                'admission_number': student.admission_number
            },
            'summary': {
                'total_days': total_days,
                'present': present,
                'absent': absent,
                'late': late,
                'half_day': half_day,
                'attendance_percentage': round(attendance_percentage, 2)
            },
            'daily_records': daily_records
        })
    
    def get_class_report(self, request):
        class_id = request.query_params.get('class_id')
        section_id = request.query_params.get('section_id')
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        
        if not class_id:
            return Response({'error': 'class_id required'}, status=400)
        
        students = Student.objects.filter(
            current_class_id=class_id,
            is_active=True
        )
        
        if section_id:
            students = students.filter(current_section_id=section_id)
        
        attendances = StudentAttendance.objects.filter(
            student__in=students
        )
        
        if month and year:
            attendances = attendances.filter(
                date__year=year,
                date__month=month
            )
        
        total_days = attendances.values('date').distinct().count()
        
        report = []
        for student in students.order_by('roll_number'):
            student_attendance = attendances.filter(student=student)
            present = student_attendance.filter(status='present').count()
            absent = student_attendance.filter(status='absent').count()
            late = student_attendance.filter(status='late').count()
            half_day = student_attendance.filter(status='half_day').count()
            
            total_present = present + late + (half_day * 0.5)
            percentage = (total_present / total_days * 100) if total_days > 0 else 0
            
            report.append({
                'student_id': student.id,
                'name': student.user.full_name,
                'roll_number': student.roll_number,
                'present': present,
                'absent': absent,
                'late': late,
                'half_day': half_day,
                'attendance_percentage': round(percentage, 2)
            })
        
        return Response({
            'class': Class.objects.get(id=class_id).name,
            'section': Section.objects.get(id=section_id).name if section_id else 'All',
            'month': month,
            'year': year,
            'total_days': total_days,
            'total_students': students.count(),
            'report': report
        })
    
    def get_monthly_report(self, request):
        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))
        
        # Get all attendance for the month
        attendances = StudentAttendance.objects.filter(
            date__year=year,
            date__month=month
        )
        
        # Daily attendance trend
        daily = attendances.values('date').annotate(
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
            half_day=Count('id', filter=Q(status='half_day'))
        ).order_by('date')
        
        # Class-wise average
        class_avg = attendances.values(
            'class_group__name',
            'section__name'
        ).annotate(
            avg_attendance=Avg(
                Case(
                    When(status='present', then=1.0),
                    When(status='late', then=1.0),
                    When(status='half_day', then=0.5),
                    default=0.0,
                    output_field=FloatField()
                )
            ) * 100,
            total_students=Count('student', distinct=True)
        ).order_by('class_group__name', 'section__name')
        
        # Top and bottom performers
        student_attendance = attendances.values(
            'student_id',
            'student__user__first_name',
            'student__user__last_name',
            'student__roll_number'
        ).annotate(
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
            half_day=Count('id', filter=Q(status='half_day')),
            total_days=Count('id')
        )
        
        for s in student_attendance:
            total_present = s['present'] + s['late'] + (s['half_day'] * 0.5)
            s['percentage'] = (total_present / s['total_days'] * 100) if s['total_days'] > 0 else 0
        
        top_10 = sorted(student_attendance, key=lambda x: x['percentage'], reverse=True)[:10]
        bottom_10 = sorted(student_attendance, key=lambda x: x['percentage'])[:10]
        
        return Response({
            'month': datetime(year, month, 1).strftime('%B %Y'),
            'daily_trend': daily,
            'class_average': class_avg,
            'top_10': top_10,
            'bottom_10': bottom_10
        })

class ExportReportView(APIView):
    """
    Export reports in various formats (CSV, PDF, Excel)
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        serializer = ExportFormatSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        format_type = data['format']
        report_type = data['report_type']
        
        if report_type == 'student':
            report_data = self.get_student_data(data.get('filters', {}))
        elif report_type == 'attendance':
            report_data = self.get_attendance_data(data.get('filters', {}))
        elif report_type == 'finance':
            report_data = self.get_finance_data(data.get('filters', {}))
        elif report_type == 'exam':
            report_data = self.get_exam_data(data.get('filters', {}))
        elif report_type == 'library':
            report_data = self.get_library_data(data.get('filters', {}))
        else:
            return Response({'error': 'Invalid report type'}, status=400)
        
        if format_type == 'csv':
            return self.export_csv(report_type, report_data)
        elif format_type == 'excel':
            return self.export_excel(report_type, report_data)
        elif format_type == 'pdf':
            return self.export_pdf(report_type, report_data)
    
    def get_student_data(self, filters):
        students = Student.objects.filter(is_active=True).select_related(
            'user', 'current_class', 'current_section'
        )
        
        if filters.get('class_id'):
            students = students.filter(current_class_id=filters['class_id'])
        if filters.get('section_id'):
            students = students.filter(current_section_id=filters['section_id'])
        if filters.get('gender'):
            students = students.filter(gender=filters['gender'])
        
        data = []
        for student in students.order_by('current_class__name', 'roll_number'):
            data.append({
                'Admission No': student.admission_number,
                'Roll No': student.roll_number,
                'Name': student.user.full_name,
                'Class': student.current_class.name if student.current_class else '',
                'Section': student.current_section.name if student.current_section else '',
                'Gender': student.get_gender_display(),
                'Date of Birth': student.date_of_birth,
                'Blood Group': student.blood_group,
                'Address': f"{student.address_line1}, {student.city}",
                'Phone': student.user.phone,
                'Email': student.user.email
            })
        
        return data
    
    def get_attendance_data(self, filters):
        start_date = filters.get('start_date')
        end_date = filters.get('end_date')
        
        if not start_date:
            start_date = timezone.now().date().replace(day=1)
        if not end_date:
            end_date = timezone.now().date()
        
        attendances = StudentAttendance.objects.filter(
            date__range=[start_date, end_date]
        ).select_related('student__user', 'class_group', 'section')
        
        if filters.get('class_id'):
            attendances = attendances.filter(class_group_id=filters['class_id'])
        if filters.get('section_id'):
            attendances = attendances.filter(section_id=filters['section_id'])
        
        data = []
        for att in attendances.order_by('date', 'student__roll_number'):
            data.append({
                'Date': att.date,
                'Student Name': att.student.user.full_name,
                'Roll No': att.student.roll_number,
                'Class': att.class_group.name,
                'Section': att.section.name,
                'Status': att.get_status_display(),
                'Remarks': att.remarks
            })
        
        return data
    
    def get_finance_data(self, filters):
        start_date = filters.get('start_date')
        end_date = filters.get('end_date')
        
        payments = Payment.objects.filter(status='completed').select_related(
            'student__user', 'fee_assignment__fee_structure', 'received_by'
        )
        
        if start_date:
            payments = payments.filter(payment_date__gte=start_date)
        if end_date:
            payments = payments.filter(payment_date__lte=end_date)
        if filters.get('payment_method'):
            payments = payments.filter(payment_method=filters['payment_method'])
        
        data = []
        for payment in payments.order_by('-payment_date'):
            data.append({
                'Receipt No': payment.receipt_number,
                'Date': payment.payment_date,
                'Student': payment.student.user.full_name,
                'Admission No': payment.student.admission_number,
                'Fee Type': payment.fee_assignment.fee_structure.name if payment.fee_assignment else 'Other',
                'Amount': payment.amount,
                'Payment Mode': payment.get_payment_method_display(),
                'Transaction ID': payment.transaction_id or '',
                'Received By': payment.received_by.full_name if payment.received_by else ''
            })
        
        return data
    
    def get_exam_data(self, filters):
        exam_id = filters.get('exam_id')
        class_id = filters.get('class_id')
        
        results = ExamResult.objects.all().select_related(
            'student__user', 'exam_schedule__exam', 'exam_schedule__subject'
        )
        
        if exam_id:
            results = results.filter(exam_schedule__exam_id=exam_id)
        if class_id:
            results = results.filter(student__current_class_id=class_id)
        
        data = []
        for result in results.order_by('student__roll_number'):
            data.append({
                'Exam': result.exam_schedule.exam.name,
                'Subject': result.exam_schedule.subject.name,
                'Student': result.student.user.full_name,
                'Roll No': result.student.roll_number,
                'Class': result.student.current_class.name,
                'Marks Obtained': result.marks_obtained,
                'Max Marks': result.exam_schedule.max_marks,
                'Percentage': f"{result.marks_obtained / result.exam_schedule.max_marks * 100:.2f}%",
                'Grade': result.grade
            })
        
        return data
    
    def get_library_data(self, filters):
        issues = BookIssue.objects.all().select_related(
            'book', 'student__user', 'teacher__user'
        )
        
        if filters.get('status'):
            issues = issues.filter(status=filters['status'])
        if filters.get('from_date'):
            issues = issues.filter(issue_date__gte=filters['from_date'])
        
        data = []
        for issue in issues.order_by('-issue_date'):
            borrower = issue.student or issue.teacher
            borrower_name = borrower.user.full_name if borrower else 'Unknown'
            
            data.append({
                'Book Title': issue.book.title,
                'Book Author': issue.book.author,
                'ISBN': issue.book.isbn,
                'Borrower': borrower_name,
                'Borrower Type': issue.issued_to,
                'Issue Date': issue.issue_date,
                'Due Date': issue.due_date,
                'Return Date': issue.return_date or '',
                'Status': issue.get_status_display(),
                'Fine': issue.fine_amount
            })
        
        return data
    
    def export_csv(self, report_type, data):
        if not data:
            return Response({'error': 'No data to export'}, status=404)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().date()}.csv"'
        
        writer = csv.DictWriter(response, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        return response
    
    def export_excel(self, report_type, data):
        if not data:
            return Response({'error': 'No data to export'}, status=404)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{report_type.capitalize()} Report"
        
        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Write data
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item[key])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().date()}.xlsx"'
        
        wb.save(response)
        return response
    
    def export_pdf(self, report_type, data):
        if not data:
            return Response({'error': 'No data to export'}, status=404)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().date()}.pdf"'
        
        # Create PDF
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title = Paragraph(f"{report_type.capitalize()} Report", styles['Title'])
        elements.append(title)
        
        # Table data
        table_data = []
        
        # Headers
        headers = list(data[0].keys())
        table_data.append([Paragraph(h, styles['Normal']) for h in headers])
        
        # Data rows
        for item in data:
            row = []
            for key in headers:
                value = str(item[key])
                if len(value) > 30:
                    value = value[:27] + '...'
                row.append(Paragraph(value, styles['Normal']))
            table_data.append(row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return response